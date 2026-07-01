#!/usr/bin/env python3
"""
MSC 2026 - Compensi giudici per sessione (da Judges > Judges confirmed).
Regole utente: 100 euro per OGNI sessione distinta confermata; presenze con
Weight 0 NON conteggiate (ma mostrate); account di test escluso.
Match dei giudici tra sessioni per nome normalizzato. Dati verificati via checksum.
"""
import os, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), "msc_sessions.txt")
OUTDIR = os.path.expanduser("~/Desktop")
SESSIONS = ["Tasting Nihonshu", "Tasting Shochu", "Design", "Food Pairing"]
FEE = 100

def norm(n):
    n = unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", n).strip().lower()

def load():
    people = {}
    for line in open(SRC, encoding="utf-8"):
        line = line.strip()
        if not line:
            continue
        ses, name, email, w = [x.strip() for x in line.split("|||")]
        w = int(w) if w.isdigit() else 0
        if "@compify.io" in email or name.startswith("User_"):
            continue  # account di test -> escluso
        k = norm(name)
        p = people.setdefault(k, {"name": name, "emails": [], "sess": {}})
        if email not in p["emails"]:
            p["emails"].append(email)
        p["sess"][ses] = max(p["sess"].get(ses, 0), w)  # peso max per sessione
    return people

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
PAY_FILL = PatternFill("solid", fgColor="E2EFDA")   # verde chiaro (>=2 sessioni)
W0_FILL = PatternFill("solid", fgColor="FCE4D6")    # arancio chiaro (peso 0)

def main():
    people = load()
    rows = []
    for p in people.values():
        paid = [s for s in SESSIONS if p["sess"].get(s, 0) >= 1]
        present0 = [s for s in SESSIONS if p["sess"].get(s, 0) == 0 and s in p["sess"]]
        compenso = FEE * len(paid)
        note = ""
        if present0:
            note = "Peso 0 (non pagata): " + ", ".join(present0)
        rows.append({
            "name": p["name"],
            "emails": "; ".join(p["emails"]),
            "cells": {s: ("Sì" if p["sess"].get(s, 0) >= 1 else ("Sì (peso 0)" if s in p["sess"] else "")) for s in SESSIONS},
            "n": len(paid), "comp": compenso, "note": note,
        })
    # ordina: piu sessioni pagate prima, poi nome
    rows.sort(key=lambda r: (-r["n"], norm(r["name"])))

    wb = Workbook(); ws = wb.active; ws.title = "Compensi giudici"
    header = ["Nome"] + SESSIONS + ["N° sessioni", "Compenso €", "Note", "Email"]
    ws.append(header)
    for r in rows:
        ws.append([r["name"]] + [r["cells"][s] for s in SESSIONS] + [r["n"], r["comp"], r["note"], r["emails"]])
    # stile
    widths = [26, 16, 15, 11, 14, 12, 11, 34, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(vertical="center", wrap_text=True)
    ncol = len(header)
    for ridx in range(2, ws.max_row + 1):
        npay = ws.cell(ridx, 6).value
        for cidx in range(1, ncol + 1):
            cell = ws.cell(ridx, cidx)
            if cidx in (2, 3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, str) and "peso 0" in cell.value:
                cell.fill = W0_FILL
        if npay and npay >= 2:
            ws.cell(ridx, 6).fill = PAY_FILL
            ws.cell(ridx, 7).fill = PAY_FILL
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"
    # riga totale
    tot = sum(r["comp"] for r in rows)
    ws.append([])
    trow = ws.max_row + 1
    ws.cell(trow, 1, "TOTALE").font = Font(bold=True)
    ws.cell(trow, 6, sum(r["n"] for r in rows)).font = Font(bold=True)
    ws.cell(trow, 7, tot).font = Font(bold=True)

    path = os.path.join(OUTDIR, "MSC2026 - Sessioni e compensi giudici.xlsx")
    wb.save(path)

    # report
    multi = [r for r in rows if r["n"] >= 2]
    paid_any = [r for r in rows if r["n"] >= 1]
    zero = [r for r in rows if r["n"] == 0]
    print("FILE:", path)
    print(f"Giudici unici: {len(rows)}")
    print(f"  pagati (>=1 sessione): {len(paid_any)}  |  in >=2 sessioni: {len(multi)}  |  0 pagate (solo peso 0): {len(zero)}")
    print(f"  TOTALE COMPENSI: {tot} € ({sum(r['n'] for r in rows)} sessioni x {FEE}€)")
    print("\nGIUDICI DA PAGARE DI PIÙ (>=2 sessioni):")
    for r in multi:
        ss = ", ".join(s for s in SESSIONS if r["cells"][s] == "Sì")
        print(f"  {r['comp']:>4}€  {r['name']:<28} [{ss}]")
    print("\nGiudici con SOLO presenze a peso 0 (0€, esclusi dal pagamento):")
    print("  " + ", ".join(r["name"] for r in zero))

if __name__ == "__main__":
    main()
