#!/usr/bin/env python3
"""
Genera 3 elenchi XLSX per MSC 2026 dai dati Compify (scraped -> scripts/msc_data.txt):
  1. Tax Free   : Nome, Cognome, Email, IBAN            (pagamenti bonifico)
  2. Fatture    : Nome e Cognome, Email                  (giudici che emettono fattura)
  3. Consegna Sake: Nome, Cognome, Indirizzo, Telefono, Indicazioni consegna
Dati verificati via checksum col browser. Duplicati esatti rimossi (segnalati).
"""
import os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = os.path.join(os.path.dirname(__file__), "msc_data.txt")
OUTDIR = os.path.expanduser("~/Desktop")

def load():
    txt = open(SRC, encoding="utf-8").read()
    def sec(a, b):
        body = txt.split(a, 1)[1].split(b, 1)[0]
        rows = []
        for line in body.split("\n"):
            line = line.strip()
            if not line:
                continue
            rows.append([c.strip() for c in line.split("|||")])
        return rows
    return sec("@@TF@@", "@@INV@@"), sec("@@INV@@", "@@GIFT@@"), sec("@@GIFT@@", "@@END@@")

def clean_iban(v):
    v = v.strip()
    v = re.sub(r'^IBAN', '', v, flags=re.I).strip()  # rimuove prefisso "IBAN"
    v = v.lstrip(':').strip()
    return v

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

def style_sheet(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def save(wb, name):
    path = os.path.join(OUTDIR, name)
    wb.save(path)
    return path

def main():
    tf, inv, gift = load()
    report = []

    # ---------- 1. TAX FREE ----------
    wb = Workbook(); ws = wb.active; ws.title = "Tax Free"
    ws.append(["Nome", "Cognome", "Email", "IBAN"])
    seen = {}; tf_rows = 0; dup_tf = []
    for idx, nome, cognome, email, iban in tf:
        ibanc = clean_iban(iban)
        key = (cognome.lower(), nome.lower(), ibanc.replace(" ", "").lower())
        if key in seen:
            dup_tf.append(f"{nome} {cognome} (riga {idx} = riga {seen[key]})")
            continue
        seen[key] = idx
        ws.append([nome, cognome, email, ibanc]); tf_rows += 1
    style_sheet(ws, [18, 22, 34, 34])
    p1 = save(wb, "MSC2026 - Tax Free (pagamenti).xlsx")
    report.append(f"Tax Free: {tf_rows} righe (duplicati rimossi: {len(dup_tf)} -> {', '.join(dup_tf)})")

    # ---------- 2. FATTURE ----------
    wb = Workbook(); ws = wb.active; ws.title = "Fatture"
    ws.append(["Nome e Cognome", "Email"])
    for idx, nominativo, email in inv:
        ws.append([nominativo, email])
    style_sheet(ws, [30, 36])
    p2 = save(wb, "MSC2026 - Fatture.xlsx")
    report.append(f"Fatture: {len(inv)} righe")

    # ---------- 3. CONSEGNA SAKE ----------
    wb = Workbook(); ws = wb.active; ws.title = "Consegna Sake"
    ws.append(["Nome", "Cognome", "Indirizzo", "Telefono", "Indicazioni consegna"])
    seen = {}; gift_rows = 0; dup_g = []
    for row in gift:
        idx, nome, cognome, indirizzo, telefono, indicazioni = (row + [""] * 6)[:6]
        key = (nome.lower(), cognome.lower(), indirizzo.lower())
        if key in seen:
            dup_g.append(f"{nome} {cognome} (riga {idx} = riga {seen[key]})")
            continue
        seen[key] = idx
        ws.append([nome, cognome, indirizzo, telefono, indicazioni]); gift_rows += 1
    style_sheet(ws, [16, 20, 46, 16, 50])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
    p3 = save(wb, "MSC2026 - Consegna Sake.xlsx")
    report.append(f"Consegna Sake: {gift_rows} righe (duplicati rimossi: {len(dup_g)} -> {', '.join(dup_g) or 'nessuno'})")

    print("FILE CREATI:")
    for p in (p1, p2, p3):
        print("  " + p)
    print("\nRIEPILOGO:")
    for r in report:
        print("  - " + r)

if __name__ == "__main__":
    main()
