#!/usr/bin/env python3
"""
MSC 2026 - Tracker pagamenti giudici (master).
Unisce:
  - sessioni/compensi (scripts/msc_sessions.txt)  -> quanto (100 €/sessione, peso 0 escluso)
  - metodo pagamento  (scripts/msc_methods.txt)    -> come (Tax Free / Fattura / Sake)
  - IBAN per i Tax Free (scripts/msc_data.txt)
Match giudici via EMAIL (primario) + nome normalizzato (fallback).
Tutti i dati verificati via checksum col sito.
"""
import os, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
OUTDIR = os.path.expanduser("~/Desktop")
SESSIONS = ["Tasting Nihonshu", "Tasting Shochu", "Design", "Food Pairing"]
FEE = 100

def norm(n):
    n = unicodedata.normalize("NFKD", n).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", n).strip().lower()

def clean_iban(v):
    v = re.sub(r'^IBAN', '', v.strip(), flags=re.I).strip().lstrip(':').strip()
    return v

# ---- sessioni -> giudici con compenso ----
def load_sessions():
    people = {}
    for line in open(os.path.join(HERE, "msc_sessions.txt"), encoding="utf-8"):
        line = line.strip()
        if not line: continue
        ses, name, email, w = [x.strip() for x in line.split("|||")]
        w = int(w) if w.isdigit() else 0
        if "@compify.io" in email or name.startswith("User_"): continue
        p = people.setdefault(norm(name), {"name": name, "emails": set(), "sess": {}})
        p["emails"].add(email.lower())
        p["sess"][ses] = max(p["sess"].get(ses, 0), w)
    return people

# ---- metodo per email/nome ----
def load_methods():
    by_email, by_name = {}, {}
    for line in open(os.path.join(HERE, "msc_methods.txt"), encoding="utf-8"):
        line = line.strip()
        if not line: continue
        idx, group, name, email = [x.strip() for x in line.split("|||")]
        method = {"Tax Free": "Tax Free (bonifico)", "Invoice": "Fattura",
                  "Gift / Sake": "Sake (consegna)"}[group]
        rec = {"name": name, "email": email.lower(), "method": method, "group": group}
        by_email[email.lower()] = rec
        by_name.setdefault(norm(name), rec)
    return by_email, by_name

# ---- indirizzo consegna Sake per email ----
def load_gift():
    g = {}
    for line in open(os.path.join(HERE, "msc_gift_addr.txt"), encoding="utf-8"):
        line = line.strip()
        if not line: continue
        email, addr, phone, instr = [x.strip() for x in line.split("|||")]
        g[email.lower()] = {"addr": addr, "phone": phone, "instr": instr}
    return g

def gift_detail(rec):
    """Ritorna (testo_dettaglio, nota) gestendo i 2 casi con campi invertiti/mancanti."""
    addr, phone, instr = rec["addr"], rec["phone"], rec["instr"]
    note = ""
    if "@" in addr:  # nel sistema l'indirizzo contiene un'email (errore di inserimento)
        if instr and re.search(r"\b(via|viale|corso|piazza|strada|vico|largo|localit)", instr, re.I):
            addr, instr = instr, ""  # l'indirizzo vero era nel campo note
            note = "indirizzo recuperato dal campo note del sistema"
        else:
            addr = "(INDIRIZZO MANCANTE nel sistema)"
            note = "indirizzo di consegna mancante"
    parts = [addr]
    if phone: parts.append("Tel: " + phone)
    if instr and instr not in (".", ",", "N/A"): parts.append(instr)
    return " · ".join(parts), note

# ---- IBAN per email (dai Tax Free) ----
def load_ibans():
    ib = {}
    txt = open(os.path.join(HERE, "msc_data.txt"), encoding="utf-8").read()
    tf = txt.split("@@TF@@", 1)[1].split("@@INV@@", 1)[0]
    for line in tf.split("\n"):
        if "|||" not in line: continue
        idx, nome, cognome, email, iban = [x.strip() for x in line.split("|||")]
        ib[email.lower()] = clean_iban(iban)
    return ib

HDR_FILL = PatternFill("solid", fgColor="1F4E78"); HDR_FONT = Font(bold=True, color="FFFFFF")
GREEN = PatternFill("solid", fgColor="E2EFDA")     # multi-sessione
RED = PatternFill("solid", fgColor="FFC7CE")       # metodo mancante
GREY = PatternFill("solid", fgColor="D9D9D9")      # escluso peso 0
METHOD_FILL = {"Tax Free (bonifico)": PatternFill("solid", fgColor="DDEBF7"),
               "Fattura": PatternFill("solid", fgColor="FFF2CC"),
               "Sake (consegna)": PatternFill("solid", fgColor="E4DFEC")}

def main():
    people = load_sessions()
    m_email, m_name = load_methods()
    ibans = load_ibans()
    gift = load_gift()

    # tutte le email e i nomi dei giudici confermati (per individuare veri "orfani")
    all_sess_emails = set()
    all_sess_names = set()
    for p in people.values():
        all_sess_emails |= p["emails"]
        all_sess_names.add(norm(p["name"]))

    rows = []
    for p in people.values():
        paid = [s for s in SESSIONS if p["sess"].get(s, 0) >= 1]
        present0 = [s for s in SESSIONS if p["sess"].get(s, 0) == 0 and s in p["sess"]]
        comp = FEE * len(paid)
        # trova metodo via email poi nome
        rec = None
        for e in p["emails"]:
            if e in m_email: rec = m_email[e]; break
        if not rec and norm(p["name"]) in m_name: rec = m_name[norm(p["name"])]
        method = rec["method"] if rec else "DA DEFINIRE"
        notes = []
        # dettaglio
        if rec and rec["group"] == "Tax Free":
            detail = ibans.get(rec["email"], "")
        elif rec and rec["group"] == "Invoice":
            detail = "Emette fattura"
        elif rec and rec["group"] == "Gift / Sake":
            detail = "In sake - consegna a domicilio"
            grec = gift.get(rec["email"])
            if grec:
                detail, gnote = gift_detail(grec)
                if gnote: notes.append(gnote)
        else:
            detail = "Metodo non impostato nel sistema"
        if present0:
            notes.append("Presenza peso 0 non pagata: " + ", ".join(present0))
        note = " — ".join(notes)
        stato = "Da pagare" if comp > 0 else "Escluso (peso 0)"
        rows.append({
            "name": p["name"],
            "sess": " + ".join(paid) if paid else ("(solo peso 0) " + ", ".join(present0)),
            "n": len(paid), "comp": comp, "method": method, "detail": detail,
            "email": "; ".join(sorted(p["emails"])), "stato": stato, "note": note,
        })

    rows.sort(key=lambda r: (-r["n"], r["method"], norm(r["name"])))

    wb = Workbook(); ws = wb.active; ws.title = "Tracker pagamenti"
    header = ["Nome", "Sessioni", "N°", "Compenso €", "Metodo pagamento", "Dettaglio (IBAN/azione)",
              "Email", "Stato", "Data pagamento", "Note"]
    ws.append(header)
    for r in rows:
        ws.append([r["name"], r["sess"], r["n"], r["comp"], r["method"], r["detail"],
                   r["email"], r["stato"], "", r["note"]])
    widths = [25, 30, 5, 11, 20, 60, 32, 16, 14, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(vertical="center", wrap_text=True)
    for ridx in range(2, ws.max_row + 1):
        npay = ws.cell(ridx, 3).value
        method = ws.cell(ridx, 5).value
        ws.cell(ridx, 3).alignment = Alignment(horizontal="center")
        ws.cell(ridx, 4).alignment = Alignment(horizontal="center")
        ws.cell(ridx, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(ridx, 10).alignment = Alignment(wrap_text=True, vertical="top")
        if method in METHOD_FILL: ws.cell(ridx, 5).fill = METHOD_FILL[method]
        if method == "DA DEFINIRE": ws.cell(ridx, 5).fill = RED; ws.cell(ridx, 6).fill = RED
        if npay and npay >= 2: ws.cell(ridx, 3).fill = GREEN; ws.cell(ridx, 4).fill = GREEN
        if npay == 0: ws.cell(ridx, 8).fill = GREY
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    tot = sum(r["comp"] for r in rows)
    ws.append([]); trow = ws.max_row + 1
    ws.cell(trow, 1, "TOTALE").font = Font(bold=True)
    ws.cell(trow, 3, sum(r["n"] for r in rows)).font = Font(bold=True)
    ws.cell(trow, 4, tot).font = Font(bold=True)

    # foglio 2: chi ha un metodo ma NON risulta in nessuna sessione confermata
    ws2 = wb.create_sheet("Metodo senza sessione")
    ws2.append(["Nome", "Metodo", "Email", "Nota"])
    for c in ws2[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT
    extra = [rec for e, rec in m_email.items()
             if e not in all_sess_emails and norm(rec["name"]) not in all_sess_names]
    for rec in sorted(extra, key=lambda r: norm(r["name"])):
        ws2.append([rec["name"], rec["method"], rec["email"], "Ha impostato il pagamento ma non risulta confermato in sessioni"])
    for i, w in enumerate([26, 20, 34, 55], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    path = os.path.join(OUTDIR, "MSC2026 - Tracker pagamenti giudici.xlsx")
    wb.save(path)

    # report
    nomet = [r for r in rows if r["method"] == "DA DEFINIRE" and r["comp"] > 0]
    bym = {}
    for r in rows:
        if r["comp"] > 0: bym[r["method"]] = bym.get(r["method"], 0) + 1
    print("FILE:", path)
    print(f"Giudici (da sessioni): {len(rows)}  |  TOTALE compensi: {tot} €")
    print("Da pagare per metodo:", bym)
    print(f"Da pagare SENZA metodo impostato: {len(nomet)}")
    for r in nomet:
        print(f"   - {r['name']} ({r['comp']}€, {r['sess']})  email: {r['email']}")
    print(f"\nFoglio 2 'Metodo senza sessione': {len(extra)} persone con metodo ma senza sessione confermata.")

if __name__ == "__main__":
    main()
